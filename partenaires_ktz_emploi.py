import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from fpdf import FPDF
import os

OUTPUT_DIR = "C:/Users/donald-chrysostome.k/rca-jobs"

# ─── DATA ────────────────────────────────────────────────────────────────────

SECTIONS = [
    {
        "titre": "PRIORITÉ IMMÉDIATE (0-3 mois)",
        "couleur": "E74C3C",  # rouge
        "partenaires": [
            {
                "num": 1,
                "nom": "Bangui Hub Foundation",
                "type": "Incubateur local",
                "financement": "Accompagnement + réseau",
                "contact": "Facebook / LinkedIn",
                "email": "www.facebook.com/banguihub",
                "lien": "banguihub.org",
            },
            {
                "num": 2,
                "nom": "FUZE / Digital Africa (AFD)",
                "type": "Fonds pré-seed",
                "financement": "20 000 – 100 000 EUR, sans dilution",
                "contact": "Formulaire en ligne",
                "email": "application.fuze.digital-africa.co",
                "lien": "digital-africa.co/en/fuze",
            },
            {
                "num": 3,
                "nom": "Ambassade de France — SCAC Bangui",
                "type": "Coopération bilatérale",
                "financement": "Subvention FSD",
                "contact": "Formulaire site officiel",
                "email": "cf.ambafrance.org/Demande-de-financement",
                "lien": "cf.ambafrance.org",
            },
        ],
    },
    {
        "titre": "COURT TERME (3-6 mois)",
        "couleur": "E67E22",  # orange
        "partenaires": [
            {
                "num": 4,
                "nom": "PNUD Centrafrique",
                "type": "Agence ONU",
                "financement": "Subvention / partenariat",
                "contact": "registry.cf@undp.org",
                "email": "registry.cf@undp.org",
                "lien": "undp.org/fr/central-african-republic",
            },
            {
                "num": 5,
                "nom": "Telecel — ASIP Program",
                "type": "Accélérateur télécom",
                "financement": "15 000 EUR cash + 500K EUR avantages tech",
                "contact": "asiprogram.com/contact-2",
                "email": "asiprogram.com/contact-2",
                "lien": "asiprogram.com",
            },
            {
                "num": 6,
                "nom": "Orange Centrafrique",
                "type": "Opérateur télécom",
                "financement": "Partenariat data / distribution 4G",
                "contact": "ocf.serviceclient@orange.com",
                "email": "ocf.serviceclient@orange.com",
                "lien": "orangerca.com/business",
            },
        ],
    },
    {
        "titre": "MOYEN TERME (6-12 mois)",
        "couleur": "27AE60",  # vert
        "partenaires": [
            {
                "num": 7,
                "nom": "Banque Mondiale — i-COMPETE / ACFPE",
                "type": "Institution financière internationale",
                "financement": "Don IDA 30M USD (via ACFPE)",
                "contact": "acfpe-rca@acfpe.info  |  (+236) 21611255",
                "email": "acfpe-rca@acfpe.info",
                "lien": "acfpe.info",
            },
            {
                "num": 8,
                "nom": "OIT / ILO — programme JPR",
                "type": "Agence ONU emploi",
                "financement": "Subvention + assistance technique",
                "contact": "ilo@ilo.org",
                "email": "ilo@ilo.org",
                "lien": "ilo.org/africa/RCA",
            },
            {
                "num": 9,
                "nom": "UNICEF Venture Fund",
                "type": "Fonds innovation ONU",
                "financement": "Jusqu'à 100 000 USD, sans dilution",
                "contact": "Formulaire site officiel",
                "email": "unicefventurefund.org",
                "lien": "unicefventurefund.org",
            },
            {
                "num": 10,
                "nom": "BAD — Fonds Entrepreneuriat Jeunesse",
                "type": "Banque de développement",
                "financement": "Subvention / cofinancement",
                "contact": "Portail financement en ligne",
                "email": "afdb.org/jobs-for-youth",
                "lien": "afdb.org/jobs-for-youth",
            },
        ],
    },
    {
        "titre": "LONG TERME — Levée de fonds equity",
        "couleur": "2980B9",  # bleu
        "partenaires": [
            {
                "num": 11,
                "nom": "Seedstars Africa Ventures",
                "type": "VC seed (BAD + BEI)",
                "financement": "250K – 2M USD (equity)",
                "contact": "Formulaire site officiel",
                "email": "seedstars.com/contact-us",
                "lien": "seedstars-africa.vc",
            },
            {
                "num": 12,
                "nom": "Ajim Capital",
                "type": "VC pre-seed Afrique",
                "financement": "100K – 250K USD (equity)",
                "contact": "Formulaire en ligne (réponse < 2 semaines)",
                "email": "ajimcapital.com/founders",
                "lien": "ajimcapital.com",
            },
            {
                "num": 13,
                "nom": "Proparco / Choose Africa",
                "type": "Agence française (AFD)",
                "financement": "500K – 5M EUR",
                "contact": "proparco@proparco.fr  |  +33 1 53 44 31 08",
                "email": "proparco@proparco.fr",
                "lien": "proparco.fr/formulaire/nous-contacter",
            },
        ],
    },
    {
        "titre": "AUTRES — Institutionnels & Régionaux",
        "couleur": "8E44AD",  # violet
        "partenaires": [
            {
                "num": 14,
                "nom": "GIZ — Bureau Yaoundé",
                "type": "Coopération allemande",
                "financement": "Subvention + assistance technique",
                "contact": "giz-kamerun@giz.de  |  +237 222 21 23 87",
                "email": "giz-kamerun@giz.de",
                "lien": "giz.de/en/worldwide/345",
            },
            {
                "num": 15,
                "nom": "USAID — YALI / Digital Africa",
                "type": "Coopération américaine",
                "financement": "Subvention entrepreneuriat digital",
                "contact": "Via bureau YALI Afrique de l'Est",
                "email": "yali.state.gov",
                "lien": "usaid.gov/youthimpact",
            },
            {
                "num": 16,
                "nom": "BDEAC",
                "type": "Banque CEMAC",
                "financement": "Prêts PME (via IFN locales)",
                "contact": "n.ngartel@bdeac.org",
                "email": "n.ngartel@bdeac.org",
                "lien": "bdeac.org",
            },
            {
                "num": 17,
                "nom": "Centre Digital — Univ. de Bangui",
                "type": "Incubateur public (BAD/UE)",
                "financement": "Formation + hébergement startup",
                "contact": "Via Rectorat Université de Bangui",
                "email": "universite-bangui.cf",
                "lien": "afdb.org — Centre Bangui",
            },
            {
                "num": 18,
                "nom": "ActivSpaces",
                "type": "Incubateur régional (Cameroun)",
                "financement": "Accompagnement + réseau régional",
                "contact": "Via site officiel",
                "email": "activspaces.com",
                "lien": "activspaces.com",
            },
            {
                "num": 19,
                "nom": "AfriLabs",
                "type": "Réseau 500+ hubs panafricains",
                "financement": "Accès Digital Africa (10K – 300K EUR)",
                "contact": "Via site officiel",
                "email": "afrilabs.com",
                "lien": "afrilabs.com",
            },
            {
                "num": 20,
                "nom": "MINUSCA",
                "type": "Mission ONU en RCA",
                "financement": "Partenariat institutionnel",
                "contact": "Via bureau MINUSCA Bangui",
                "email": "minusca.unmissions.org",
                "lien": "minusca.unmissions.org",
            },
        ],
    },
]

# ─── EXCEL ───────────────────────────────────────────────────────────────────

def hex_to_rgb(h):
    h = h.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def make_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Partenaires KTZ Emploi"

    # ── Titre principal ──────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws["A1"] = "PARTENAIRES POTENTIELS — KTZ EMPLOI / RCA-JOBS"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill("solid", fgColor="F97316")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:H2")
    ws["A2"] = "Application mobile Android d'offres d'emploi en République Centrafricaine"
    ws["A2"].font = Font(italic=True, size=11, color="555555", name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # ── En-têtes colonnes ────────────────────────────────────────────────────
    headers = ["#", "Partenaire", "Type", "Financement", "Contact", "Email / Formulaire", "Site / Lien", "Priorité"]
    col_widths = [4, 30, 22, 38, 28, 38, 36, 20]

    header_row = 4
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor="2C3E50")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="FFFFFF")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[header_row].height = 28

    # ── Données ──────────────────────────────────────────────────────────────
    current_row = header_row + 1
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for section in SECTIONS:
        couleur = section["couleur"]
        titre   = section["titre"]

        # Ligne de section
        ws.merge_cells(f"A{current_row}:H{current_row}")
        sc = ws.cell(row=current_row, column=1, value=f"  {titre}")
        sc.font  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        sc.fill  = PatternFill("solid", fgColor=couleur)
        sc.alignment = Alignment(vertical="center")
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for p in section["partenaires"]:
            row_data = [
                p["num"],
                p["nom"],
                p["type"],
                p["financement"],
                p["contact"],
                p["email"],
                p["lien"],
                titre.split("(")[0].strip(),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.font      = Font(name="Calibri", size=9)
                cell.border    = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                # Alternance légère
                if current_row % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F9F9F9")
                else:
                    cell.fill = PatternFill("solid", fgColor="FFFFFF")
                # Colonne numéro : centré + gras
                if col == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                    cell.font = Font(bold=True, name="Calibri", size=9, color=couleur)
                # Colonne nom : gras
                if col == 2:
                    cell.font = Font(bold=True, name="Calibri", size=9)
            ws.row_dimensions[current_row].height = 42
            current_row += 1

        current_row += 1  # espace entre sections

    # ── Figer les lignes d'en-tête ───────────────────────────────────────────
    ws.freeze_panes = f"A{header_row + 1}"

    # ── Note de bas de page ──────────────────────────────────────────────────
    ws.cell(row=current_row, column=1, value="Généré le 12/05/2026 — KTZ Emploi / rca-jobs")
    ws.cell(row=current_row, column=1).font = Font(italic=True, size=8, color="999999")
    ws.merge_cells(f"A{current_row}:H{current_row}")

    path = os.path.join(OUTPUT_DIR, "partenaires_ktz_emploi.xlsx")
    wb.save(path)
    print(f"Excel sauvegardé : {path}")
    return path

# ─── PDF ─────────────────────────────────────────────────────────────────────

def clean(text):
    """Remplace les caractères non-latin1 par des équivalents ASCII."""
    replacements = {
        "\u2014": "-", "\u2013": "-", "\u2019": "'", "\u2018": "'",
        "\u201c": '"', "\u201d": '"', "\u00e9": "e", "\u00e8": "e",
        "\u00ea": "e", "\u00eb": "e", "\u00e0": "a", "\u00e2": "a",
        "\u00e4": "a", "\u00ee": "i", "\u00ef": "i", "\u00f4": "o",
        "\u00f9": "u", "\u00fb": "u", "\u00fc": "u", "\u00e7": "c",
        "\u00c9": "E", "\u00c8": "E", "\u00c0": "A", "\u00d4": "O",
        "\u00c7": "C", "\u2026": "...", "\u00ab": "\"", "\u00bb": "\"",
        "\u00b0": " deg", "\u20ac": "EUR",
    }
    for orig, repl in replacements.items():
        text = text.replace(orig, repl)
    return text


class PDF(FPDF):
    def header(self):
        self.set_fill_color(249, 115, 22)   # orange F97316
        self.rect(0, 0, 297, 22, "F")
        self.set_font("Helvetica", "B", 13)
        self.set_text_color(255, 255, 255)
        self.set_y(6)
        self.cell(0, 10, "PARTENAIRES POTENTIELS - KTZ EMPLOI / RCA-JOBS", align="C")
        self.ln(18)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "I", 7)
        self.set_text_color(150, 150, 150)
        self.cell(0, 8, f"KTZ Emploi / rca-jobs - Genere le 12/05/2026 - Page {self.page_no()}", align="C")


def make_pdf():
    pdf = PDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Sous-titre
    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5, clean("Application mobile Android d'offres d'emploi en Republique Centrafricaine"), align="C")
    pdf.ln(6)

    # Largeurs colonnes (paysage A4 = 277mm utilisable)
    col_w = [8, 42, 30, 52, 38, 52, 40, 0]
    # La dernière colonne prend le reste
    col_w[-1] = 277 - sum(col_w[:-1])

    headers = ["#", "Partenaire", "Type", "Financement", "Contact", "Email / Formulaire", "Site / Lien", "Priorite"]
    ROW_H   = 7
    HEAD_H  = 8

    def draw_section_header(titre, couleur_hex):
        r, g, b = hex_to_rgb(couleur_hex)
        pdf.set_fill_color(r, g, b)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(sum(col_w), 7, clean(f"  {titre}"), fill=True, ln=True)
        pdf.set_text_color(0, 0, 0)

    def draw_col_headers():
        pdf.set_fill_color(44, 62, 80)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 7.5)
        for h, w in zip(headers, col_w):
            pdf.cell(w, HEAD_H, h, border=1, align="C", fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)

    draw_col_headers()

    def draw_row(p, section_titre, couleur_hex, even):
        r, g, b = hex_to_rgb(couleur_hex)
        row_data = [
            str(p["num"]),
            clean(p["nom"]),
            clean(p["type"]),
            clean(p["financement"]),
            clean(p["contact"]),
            clean(p["email"]),
            clean(p["lien"]),
            clean(section_titre.split("(")[0].strip()),
        ]

        # Calcul hauteur multi-ligne
        pdf.set_font("Helvetica", "", 7)
        max_lines = 1
        for val, w in zip(row_data, col_w):
            lines = pdf.get_string_width(val) / (w - 2) + 1
            if int(lines) > max_lines:
                max_lines = int(lines)
        row_h = max(ROW_H, ROW_H * min(max_lines, 3))

        # Fond alterné
        if even:
            pdf.set_fill_color(249, 249, 249)
        else:
            pdf.set_fill_color(255, 255, 255)

        x_start = pdf.get_x()
        y_start = pdf.get_y()

        # Check page break
        if y_start + row_h > pdf.h - 18:
            pdf.add_page()
            draw_col_headers()
            x_start = pdf.get_x()
            y_start = pdf.get_y()

        for i, (val, w) in enumerate(zip(row_data, col_w)):
            pdf.set_xy(x_start, y_start)
            if i == 0:
                pdf.set_font("Helvetica", "B", 7)
                pdf.set_text_color(r, g, b)
            elif i == 1:
                pdf.set_font("Helvetica", "B", 7)
                pdf.set_text_color(0, 0, 0)
            else:
                pdf.set_font("Helvetica", "", 7)
                pdf.set_text_color(50, 50, 50)

            pdf.multi_cell(w, ROW_H, val, border=1, fill=(i == 0 or even), max_line_height=ROW_H)
            x_start += w

        pdf.set_xy(pdf.l_margin, y_start + row_h)
        pdf.set_text_color(0, 0, 0)

    even = False
    for section in SECTIONS:
        draw_section_header(section["titre"], section["couleur"])
        for p in section["partenaires"]:
            draw_row(p, section["titre"], section["couleur"], even)
            even = not even
        pdf.ln(2)

    path = os.path.join(OUTPUT_DIR, "partenaires_ktz_emploi.pdf")
    pdf.output(path)
    print(f"PDF sauvegardé : {path}")
    return path


if __name__ == "__main__":
    make_excel()
    make_pdf()
    print("Done.")
